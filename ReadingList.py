import tkinter as tk
import openpyxl


class ReadingList:

    def __init__(self, master):
        frame = tk.Frame(master)
        frame.pack()

        wb = openpyxl.load_workbook('records.xlsx')
        sheet = wb[wb.sheetnames[0]]
        recordCount = sheet.max_row

        listBox = tk.Listbox(frame, selectmode=tk.SINGLE)

        for i in range(1, recordCount):
            listBox.insert(tk.END, sheet.cell(row=i+1, column=2).value)
            
        listBox.pack()
        
        self.btnAdd = tk.Button(frame, text="Add Record", command=self.addRecord)
        self.btnAdd.pack()

        self.btnDelete = tk.Button(frame, text="Delete Record", command=self.deleteRecord)
        self.btnDelete.pack()

        self.btnEdit = tk.Button(frame, text="Edit Record", command=self.editRecord)
        self.btnEdit.pack()
        
    def addRecord(self):
        print("")


    def deleteRecord(self):
        print("")


    def editRecord(self):
        print("")

        

        
root = tk.Tk()
b = ReadingList(root)
root.mainloop()
